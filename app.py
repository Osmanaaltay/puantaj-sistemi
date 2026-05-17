from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import sqlite3
import os
import shutil
import io
from datetime import datetime, date
import calendar
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder='.')
CORS(app)

DB_PATH = 'puantaj.db'
BACKUP_DIR = 'yedekler'
AYLIK_STANDART_SAAT = 225.0
TAVAN_SAAT = 7.5

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS personel (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ad TEXT NOT NULL, soyad TEXT NOT NULL, tc TEXT UNIQUE NOT NULL,
        aktif INTEGER DEFAULT 1, olusturma_tarihi TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS maas_gecmis (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        personel_id INTEGER NOT NULL, yil INTEGER NOT NULL, ay INTEGER,
        sabit_maas REAL NOT NULL, asgari_ucret REAL DEFAULT 0,
        UNIQUE(personel_id, yil, ay),
        FOREIGN KEY(personel_id) REFERENCES personel(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS tatil (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tarih TEXT NOT NULL UNIQUE, ad TEXT NOT NULL, tur TEXT DEFAULT 'tam'
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS puantaj (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        personel_id INTEGER NOT NULL, tarih TEXT NOT NULL,
        gun_turu TEXT DEFAULT 'normal', calisma_saati REAL,
        notlar TEXT,
        UNIQUE(personel_id, tarih),
        FOREIGN KEY(personel_id) REFERENCES personel(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS avans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        personel_id INTEGER NOT NULL, tutar REAL NOT NULL,
        tarih TEXT NOT NULL, aciklama TEXT,
        FOREIGN KEY(personel_id) REFERENCES personel(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS degisiklik_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tablo TEXT NOT NULL,
        kayit_id INTEGER,
        islem TEXT NOT NULL,
        eski_deger TEXT,
        yeni_deger TEXT,
        tarih TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS ayarlar (
        anahtar TEXT PRIMARY KEY,
        deger TEXT NOT NULL
    )''')
    conn.commit()
    conn.close()

def get_maas(personel_id, yil, ay):
    conn = get_db()
    c = conn.cursor()
    r = c.execute('SELECT * FROM maas_gecmis WHERE personel_id=? AND yil=? AND ay=?',
                  (personel_id, yil, ay)).fetchone()
    if not r:
        r = c.execute('SELECT * FROM maas_gecmis WHERE personel_id=? AND yil=? AND ay IS NULL',
                      (personel_id, yil)).fetchone()
    if not r:
        r = c.execute('SELECT * FROM maas_gecmis WHERE personel_id=? ORDER BY yil DESC, COALESCE(ay,0) DESC LIMIT 1',
                      (personel_id,)).fetchone()
    conn.close()
    return {'sabit_maas': r['sabit_maas'], 'asgari_ucret': r['asgari_ucret']} if r else {'sabit_maas': 0, 'asgari_ucret': 0}

def log_yaz(conn, tablo, kayit_id, islem, eski=None, yeni=None):
    import json
    yerel_zaman = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute(
        'INSERT INTO degisiklik_log (tablo, kayit_id, islem, eski_deger, yeni_deger, tarih) VALUES (?,?,?,?,?,?)',
        (tablo, kayit_id, islem,
         json.dumps(eski, ensure_ascii=False) if eski else None,
         json.dumps(yeni, ensure_ascii=False) if yeni else None,
         yerel_zaman)
    )

def get_tatiller_dict(yil, ay):
    conn = get_db()
    ay_basi = f"{yil}-{ay:02d}-01"
    ay_sonu = f"{yil}-{ay:02d}-{calendar.monthrange(yil, ay)[1]:02d}"
    rows = conn.execute('SELECT * FROM tatil WHERE tarih>=? AND tarih<=? ORDER BY tarih',
                        (ay_basi, ay_sonu)).fetchall()
    conn.close()
    return {r['tarih']: dict(r) for r in rows}

def hesapla_maas(personel_id, yil, ay, global_asgari=None):
    """
    Hesaplama mantığı:
    - Saatlik ücret = sabit_maas / 225
    - Normal gün: girilen saat normal saate eklenir, 7.5 üstü fazla mesai
    - Pazar günü: 0 → 7.5 normal | >0 → o kadar normal (7.5 üstü fazla mesai)
    - Tatil tam (7.5s baz): girilen saat → min(saat,7.5) normal + min(saat,7.5) fazla + max(0,saat-7.5) fazla
      yani 7.5 yazılırsa: 7.5 normal + 7.5 fazla (toplamda 7.5*2 etkisi)
    - Tatil yarım (3.75s baz): girilen saat → min(saat,3.75) normal + min(saat,3.75) fazla + max(0,saat-3.75) fazla
      yani 3.75 yazılırsa: 3.75 normal + 3.75 fazla → toplam 7.5 saat etkisi
    - Yıllık izin: 7.5 normal sayılır
    - Rapor/ücretsiz izin: o gün sayılmaz
    - Toplam ücret = (normal_saat + yillik_izin_saat) * saatlik_ucret + fazla_mesai_saat * saatlik_ucret * 1.5
    - Banka (asgari) tutarı: (global_asgari / 225) × ayın_standart_saati
      ayın_standart_saati = o aydaki Pazartesi–Cumartesi gün sayısı × 7.5
    """
    conn = get_db()
    c = conn.cursor()
    p = c.execute('SELECT * FROM personel WHERE id=?', (personel_id,)).fetchone()
    if not p:
        conn.close()
        return None

    maas_bilgi = get_maas(personel_id, yil, ay)
    sabit_maas = maas_bilgi['sabit_maas']
    saatlik_ucret = sabit_maas / AYLIK_STANDART_SAAT if sabit_maas > 0 else 0

    ay_basi = f"{yil}-{ay:02d}-01"
    ay_sonu = f"{yil}-{ay:02d}-{calendar.monthrange(yil, ay)[1]:02d}"
    kayitlar = c.execute('SELECT * FROM puantaj WHERE personel_id=? AND tarih>=? AND tarih<=?',
                         (personel_id, ay_basi, ay_sonu)).fetchall()
    avanslar = c.execute('SELECT SUM(tutar) as toplam FROM avans WHERE personel_id=? AND tarih>=? AND tarih<=?',
                         (personel_id, ay_basi, ay_sonu)).fetchone()
    conn.close()

    tatiller = get_tatiller_dict(yil, ay)

    normal_saat = 0.0
    fazla_mesai_saat = 0.0
    yillik_izin_saat = 0.0

    for k in kayitlar:
        tarih = k['tarih']
        gt = k['gun_turu']
        saat = k['calisma_saati']  # None = boş girilmemiş
        gun_tarihi = date(yil, ay, int(tarih.split('-')[2]))
        is_pazar = gun_tarihi.weekday() == 6

        # Pazar öncelikli — tatil ile çakışsa bile pazar kuralı geçerli
        if is_pazar or gt == 'pazar':
            if saat is None:
                pass  # Boş = sayılmaz
            elif saat == 0:
                normal_saat += TAVAN_SAAT  # 0 → 7.5 normal
            else:
                normal_saat += min(saat, TAVAN_SAAT)
                if saat > TAVAN_SAAT:
                    fazla_mesai_saat += saat - TAVAN_SAAT
            continue

        # Tatil günleri (pazar değilse)
        if tarih in tatiller:
            t_tur = tatiller[tarih]['tur']
            if t_tur == 'tam':
                # Boş → ekleme yok
                # 7.5 → 7.5 normal
                # >7.5 → 7.5 normal + fazlası fazla mesai
                if saat is None:
                    pass
                else:
                    normal_saat += TAVAN_SAAT
                    if saat > TAVAN_SAAT:
                        fazla_mesai_saat += saat - TAVAN_SAAT
            else:  # yarım gün (baz=3.75)
                # Boş → ekleme yok
                # 3.75 → 7.5 normal (3.75 yazılsa da tam gün sayılır)
                # >3.75 → 7.5 normal + fazlası fazla mesai
                if saat is None:
                    pass
                else:
                    normal_saat += TAVAN_SAAT  # her zaman 7.5 normal
                    if saat > TAVAN_SAAT / 2:
                        fazla_mesai_saat += saat - TAVAN_SAAT / 2
            continue

        # Rapor / ücretsiz izin → sayılmaz
        if gt in ('rapor', 'ucretsiz_izin'):
            continue

        # Yıllık izin
        if gt == 'yillik_izin':
            yillik_izin_saat += TAVAN_SAAT
            continue

        # Normal gün
        if saat is None:
            continue
        normal_saat += min(saat, TAVAN_SAAT)
        if saat > TAVAN_SAAT:
            fazla_mesai_saat += saat - TAVAN_SAAT

    # Toplam ödenen saat
    odenen_normal = normal_saat + yillik_izin_saat
    # Ücret hesabı: çalışılan × saatlik + fazla mesai × 1.5
    normal_ucret = odenen_normal * saatlik_ucret
    fazla_mesai_ucreti = fazla_mesai_saat * saatlik_ucret * 1.5
    toplam_maas = normal_ucret + fazla_mesai_ucreti

    # Banka (asgari) tutarı: personelin fiili normal_saat × (asgari / 225)
    if global_asgari is not None and global_asgari > 0:
        bankaya_giden_tutar = round(normal_saat * (global_asgari / AYLIK_STANDART_SAAT), 2)
    else:
        bankaya_giden_tutar = maas_bilgi['asgari_ucret']

    avans_toplam = avanslar['toplam'] or 0
    elden_giden = max(0, toplam_maas - bankaya_giden_tutar - avans_toplam)

    return {
        'personel': dict(p),
        'maas_bilgi': maas_bilgi,
        'saatlik_ucret': round(saatlik_ucret, 4),
        'normal_saat': round(normal_saat, 2),
        'yillik_izin_saat': round(yillik_izin_saat, 2),
        'odenen_normal': round(odenen_normal, 2),
        'fazla_mesai_saat': round(fazla_mesai_saat, 2),
        'normal_ucret': round(normal_ucret, 2),
        'fazla_mesai_ucreti': round(fazla_mesai_ucreti, 2),
        'toplam_maas': round(toplam_maas, 2),
        'avans': round(avans_toplam, 2),
        'bankaya_giden': bankaya_giden_tutar,
        'elden_giden': round(elden_giden, 2),
    }

# ── ROUTES ──────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

# Personel
@app.route('/api/personel', methods=['GET'])
def personel_listesi():
    conn = get_db()
    rows = conn.execute('SELECT * FROM personel WHERE aktif=1 ORDER BY ad').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/personel', methods=['POST'])
def personel_ekle():
    d = request.json
    conn = get_db()
    try:
        # Aktif kayıtta TC var mı?
        aktif_mevcut = conn.execute(
            'SELECT * FROM personel WHERE tc=? AND aktif=1', (d['tc'],)
        ).fetchone()
        if aktif_mevcut:
            conn.close()
            return jsonify({
                'ok': False,
                'cakisma': True,
                'mevcut': {'id': aktif_mevcut['id'], 'ad': aktif_mevcut['ad'], 'soyad': aktif_mevcut['soyad'], 'tc': aktif_mevcut['tc']}
            }), 409
        # Pasif kayıtta TC var mı? — reaktif
        pasif_mevcut = conn.execute('SELECT * FROM personel WHERE tc=? AND aktif=0', (d['tc'],)).fetchone()
        if pasif_mevcut:
            conn.execute('UPDATE personel SET aktif=1, ad=?, soyad=? WHERE id=?',
                         (d['ad'], d['soyad'], pasif_mevcut['id']))
            conn.commit()
            conn.close()
            return jsonify({'ok': True, 'id': pasif_mevcut['id'], 'reaktif': True})
        cur = conn.execute('INSERT INTO personel (ad, soyad, tc) VALUES (?,?,?)',
                           (d['ad'], d['soyad'], d['tc']))
        pid = cur.lastrowid
        log_yaz(conn, 'personel', pid, 'INSERT', yeni={'ad': d['ad'], 'soyad': d['soyad'], 'tc': d['tc']})
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'id': pid})
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'hata': str(e)}), 400

@app.route('/api/personel/<int:pid>', methods=['PUT'])
def personel_guncelle(pid):
    d = request.json
    conn = get_db()
    try:
        eski = dict(conn.execute('SELECT * FROM personel WHERE id=?', (pid,)).fetchone() or {})
        # TC çakışma kontrolü — başka aktif kişide aynı TC var mı?
        cakisan = conn.execute(
            'SELECT id, ad, soyad FROM personel WHERE tc=? AND aktif=1 AND id!=?',
            (d['tc'], pid)
        ).fetchone()
        if cakisan:
            conn.close()
            return jsonify({'ok': False, 'hata': f"Bu TC zaten kayıtlı: {cakisan['ad']} {cakisan['soyad']}"}), 409
        conn.execute('UPDATE personel SET ad=?, soyad=?, tc=? WHERE id=?',
                     (d['ad'], d['soyad'], d['tc'], pid))
        log_yaz(conn, 'personel', pid, 'UPDATE', eski=eski, yeni={'ad': d['ad'], 'soyad': d['soyad'], 'tc': d['tc']})
        if d.get('sabit_maas'):
            yil = d.get('yil', datetime.now().year)
            ay = d.get('ay') or None
            conn.execute('''INSERT INTO maas_gecmis (personel_id, yil, ay, sabit_maas, asgari_ucret)
                            VALUES (?,?,?,?,?)
                            ON CONFLICT(personel_id, yil, ay) DO UPDATE SET
                            sabit_maas=excluded.sabit_maas, asgari_ucret=excluded.asgari_ucret''',
                         (pid, yil, ay, d['sabit_maas'], d.get('asgari_ucret', 0)))
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'hata': str(e)}), 400

@app.route('/api/personel/<int:pid>', methods=['DELETE'])
def personel_sil(pid):
    conn = get_db()
    eski = dict(conn.execute('SELECT * FROM personel WHERE id=?', (pid,)).fetchone() or {})
    # İlgili tüm verileri sil
    conn.execute('DELETE FROM puantaj WHERE personel_id=?', (pid,))
    conn.execute('DELETE FROM avans WHERE personel_id=?', (pid,))
    conn.execute('DELETE FROM maas_gecmis WHERE personel_id=?', (pid,))
    conn.execute('DELETE FROM personel WHERE id=?', (pid,))
    log_yaz(conn, 'personel', pid, 'KALICI_SIL', eski=eski)
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/personel/<int:pid>/pasife_al', methods=['POST'])
def personel_pasife_al(pid):
    conn = get_db()
    eski = dict(conn.execute('SELECT * FROM personel WHERE id=?', (pid,)).fetchone() or {})
    conn.execute('UPDATE personel SET aktif=0 WHERE id=?', (pid,))
    log_yaz(conn, 'personel', pid, 'DELETE', eski=eski)
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/personel/pasif', methods=['GET'])
def pasif_personel_listesi():
    conn = get_db()
    rows = conn.execute('SELECT * FROM personel WHERE aktif=0 ORDER BY ad').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/personel/<int:pid>/aktife_al', methods=['POST'])
def personel_aktife_al(pid):
    conn = get_db()
    p = conn.execute('SELECT * FROM personel WHERE id=?', (pid,)).fetchone()
    if not p:
        conn.close()
        return jsonify({'ok': False, 'hata': 'Personel bulunamadı'}), 404
    # Aktif kayıtta TC çakışması?
    cakisan = conn.execute('SELECT id FROM personel WHERE tc=? AND aktif=1 AND id!=?', (p['tc'], pid)).fetchone()
    if cakisan:
        conn.close()
        return jsonify({'ok': False, 'hata': 'Bu TC ile aktif bir kayıt mevcut'}), 409
    conn.execute('UPDATE personel SET aktif=1 WHERE id=?', (pid,))
    log_yaz(conn, 'personel', pid, 'UPDATE', yeni={'aktif': 1})
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# Maaş geçmişi
@app.route('/api/maas/<int:pid>', methods=['GET'])
def maas_listesi(pid):
    conn = get_db()
    rows = conn.execute('SELECT * FROM maas_gecmis WHERE personel_id=? ORDER BY yil DESC, COALESCE(ay,0) DESC', (pid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/maas', methods=['POST'])
def maas_ekle():
    d = request.json
    conn = get_db()
    try:
        conn.execute('''INSERT INTO maas_gecmis (personel_id, yil, ay, sabit_maas, asgari_ucret) VALUES (?,?,?,?,?)
                        ON CONFLICT(personel_id, yil, ay) DO UPDATE SET sabit_maas=excluded.sabit_maas, asgari_ucret=excluded.asgari_ucret''',
                     (d['personel_id'], d['yil'], d.get('ay') or None, d['sabit_maas'], d.get('asgari_ucret', 0)))
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'hata': str(e)}), 400

@app.route('/api/maas/<int:maas_id>', methods=['DELETE'])
def maas_sil(maas_id):
    conn = get_db()
    eski_r = conn.execute('SELECT * FROM maas_gecmis WHERE id=?', (maas_id,)).fetchone()
    if eski_r:
        log_yaz(conn, 'maas_gecmis', maas_id, 'DELETE', eski=dict(eski_r))
    conn.execute('DELETE FROM maas_gecmis WHERE id=?', (maas_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/maas/<int:maas_id>', methods=['PUT'])
def maas_guncelle(maas_id):
    d = request.json
    conn = get_db()
    try:
        eski_r = conn.execute('SELECT * FROM maas_gecmis WHERE id=?', (maas_id,)).fetchone()
        eski = dict(eski_r) if eski_r else None
        conn.execute('UPDATE maas_gecmis SET yil=?, ay=?, sabit_maas=?, asgari_ucret=0 WHERE id=?',
                     (d['yil'], d.get('ay') or None, d['sabit_maas'], maas_id))
        log_yaz(conn, 'maas_gecmis', maas_id, 'UPDATE', eski=eski,
                yeni={'yil': d['yil'], 'ay': d.get('ay'), 'sabit_maas': d['sabit_maas']})
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'hata': str(e)}), 400

# Tatil
@app.route('/api/tatil', methods=['GET'])
def tatil_listesi():
    yil = request.args.get('yil')
    conn = get_db()
    if yil:
        rows = conn.execute("SELECT * FROM tatil WHERE tarih LIKE ? ORDER BY tarih", (f"{yil}-%",)).fetchall()
    else:
        rows = conn.execute('SELECT * FROM tatil ORDER BY tarih').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/tatil', methods=['POST'])
def tatil_ekle():
    d = request.json
    conn = get_db()
    try:
        eski_r = conn.execute('SELECT * FROM tatil WHERE tarih=?', (d['tarih'],)).fetchone()
        eski = dict(eski_r) if eski_r else None
        conn.execute('''INSERT INTO tatil (tarih, ad, tur) VALUES (?,?,?)
                        ON CONFLICT(tarih) DO UPDATE SET ad=excluded.ad, tur=excluded.tur''',
                     (d['tarih'], d['ad'], d.get('tur', 'tam')))
        yeni_r = conn.execute('SELECT * FROM tatil WHERE tarih=?', (d['tarih'],)).fetchone()
        log_yaz(conn, 'tatil', yeni_r['id'] if yeni_r else None,
                'UPDATE' if eski else 'INSERT', eski=eski,
                yeni={'tarih': d['tarih'], 'ad': d['ad'], 'tur': d.get('tur', 'tam')})
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'hata': str(e)}), 400

@app.route('/api/tatil/<int:tid>', methods=['DELETE'])
def tatil_sil(tid):
    conn = get_db()
    eski_r = conn.execute('SELECT * FROM tatil WHERE id=?', (tid,)).fetchone()
    if eski_r:
        log_yaz(conn, 'tatil', tid, 'DELETE', eski=dict(eski_r))
    conn.execute('DELETE FROM tatil WHERE id=?', (tid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# Puantaj
@app.route('/api/puantaj/<int:pid>', methods=['GET'])
def puantaj_listesi(pid):
    yil = request.args.get('yil', datetime.now().year)
    ay = request.args.get('ay', datetime.now().month)
    ay_basi = f"{yil}-{int(ay):02d}-01"
    ay_sonu = f"{yil}-{int(ay):02d}-{calendar.monthrange(int(yil), int(ay))[1]:02d}"
    conn = get_db()
    rows = conn.execute('SELECT * FROM puantaj WHERE personel_id=? AND tarih>=? AND tarih<=? ORDER BY tarih',
                        (pid, ay_basi, ay_sonu)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/puantaj', methods=['POST'])
def puantaj_kaydet():
    d = request.json
    conn = get_db()
    try:
        saat = d.get('calisma_saati')
        eski_r = conn.execute('SELECT * FROM puantaj WHERE personel_id=? AND tarih=?',
                              (d['personel_id'], d['tarih'])).fetchone()
        eski = dict(eski_r) if eski_r else None
        conn.execute('''INSERT INTO puantaj (personel_id, tarih, gun_turu, calisma_saati, notlar)
                        VALUES (?,?,?,?,?)
                        ON CONFLICT(personel_id, tarih) DO UPDATE SET
                        gun_turu=excluded.gun_turu, calisma_saati=excluded.calisma_saati, notlar=excluded.notlar''',
                     (d['personel_id'], d['tarih'], d['gun_turu'], saat, d.get('notlar', '')))
        yeni_r = conn.execute('SELECT * FROM puantaj WHERE personel_id=? AND tarih=?',
                              (d['personel_id'], d['tarih'])).fetchone()
        log_yaz(conn, 'puantaj', yeni_r['id'] if yeni_r else None,
                'UPDATE' if eski else 'INSERT', eski=eski,
                yeni={'personel_id': d['personel_id'], 'tarih': d['tarih'],
                      'gun_turu': d['gun_turu'], 'calisma_saati': saat, 'notlar': d.get('notlar', '')})
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'hata': str(e)}), 400

@app.route('/api/puantaj/toplu', methods=['POST'])
def puantaj_toplu():
    d = request.json
    pid = d['personel_id']
    kayitlar = d['kayitlar']
    conn = get_db()
    try:
        # O ay tatilleri önceden çek
        if kayitlar:
            ilk_tarih = kayitlar[0]['tarih']
            yil_t = int(ilk_tarih[:4]); ay_t = int(ilk_tarih[5:7])
            tatiller_map = get_tatiller_dict(yil_t, ay_t)
        else:
            tatiller_map = {}

        import json as _json
        degisen = []   # geri alma için: sadece değişenler
        silinen = []

        for k in kayitlar:
            tarih = k['tarih']
            saat = k.get('saat')
            eski_r = conn.execute('SELECT * FROM puantaj WHERE personel_id=? AND tarih=?', (pid, tarih)).fetchone()
            eski = dict(eski_r) if eski_r else None

            if saat is None:
                # Hücre boşaltıldı → sadece DB'de kayıt varsa sil ve logla
                if eski:
                    silinen.append(eski)
                    conn.execute('DELETE FROM puantaj WHERE personel_id=? AND tarih=?', (pid, tarih))
                continue

            gun = date(int(tarih[:4]), int(tarih[5:7]), int(tarih[8:10]))
            if gun.weekday() == 6:
                gt = 'pazar'
            elif tarih in tatiller_map:
                gt = 'tatil_tam' if tatiller_map[tarih]['tur'] == 'tam' else 'tatil_yarim'
            else:
                gt = 'normal'

            # Sadece değer gerçekten değiştiyse logla
            eski_saat = eski['calisma_saati'] if eski else None
            if eski_saat != saat:
                degisen.append({
                    'tarih': tarih,
                    'eski_saat': eski_saat,
                    'yeni_saat': saat,
                    'gun_turu': gt,
                    'eski_gun_turu': eski['gun_turu'] if eski else None,
                })

            conn.execute('''INSERT INTO puantaj (personel_id, tarih, gun_turu, calisma_saati)
                            VALUES (?,?,?,?)
                            ON CONFLICT(personel_id, tarih) DO UPDATE SET
                            gun_turu=excluded.gun_turu, calisma_saati=excluded.calisma_saati''',
                         (pid, tarih, gt, saat))

        # Tüm değişikliği tek log satırına yaz — eski değerleri de sakla (geri alma için)
        if degisen or silinen:
            p_r = conn.execute('SELECT ad, soyad FROM personel WHERE id=?', (pid,)).fetchone()
            p_ad = f"{p_r['ad']} {p_r['soyad']}" if p_r else str(pid)
            ozet = {
                'personel': p_ad,
                'personel_id': pid,
                'degistirilen_gun': len(degisen),
                'silinen_gun': len(silinen),
                'detay': degisen,
                'silinen': silinen,
            }
            conn.execute(
                'INSERT INTO degisiklik_log (tablo, kayit_id, islem, eski_deger, yeni_deger, tarih) VALUES (?,?,?,?,?,?)',
                ('puantaj', pid, 'TOPLU', None, _json.dumps(ozet, ensure_ascii=False),
                 datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            )
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'hata': str(e)}), 400

@app.route('/api/puantaj/toplu_hepsi', methods=['POST'])
def puantaj_toplu_hepsi():
    import json as _json
    d = request.json
    tum_personeller = d['personeller']  # [{personel_id, kayitlar: [{tarih, saat}]}]
    conn = get_db()
    try:
        tum_degisen = []   # tüm personellerin değişiklikleri
        tum_silinen = []

        for grup in tum_personeller:
            pid = grup['personel_id']
            kayitlar = grup['kayitlar']
            if not kayitlar:
                continue

            ilk_tarih = kayitlar[0]['tarih']
            yil_t = int(ilk_tarih[:4]); ay_t = int(ilk_tarih[5:7])
            tatiller_map = get_tatiller_dict(yil_t, ay_t)

            p_r = conn.execute('SELECT ad, soyad FROM personel WHERE id=?', (pid,)).fetchone()
            p_ad = f"{p_r['ad']} {p_r['soyad']}" if p_r else str(pid)

            for k in kayitlar:
                tarih = k['tarih']
                saat = k.get('saat')
                eski_r = conn.execute('SELECT * FROM puantaj WHERE personel_id=? AND tarih=?', (pid, tarih)).fetchone()
                eski = dict(eski_r) if eski_r else None

                if saat is None:
                    if eski:
                        tum_silinen.append({**eski, 'personel': p_ad})
                        conn.execute('DELETE FROM puantaj WHERE personel_id=? AND tarih=?', (pid, tarih))
                    continue

                gun = date(int(tarih[:4]), int(tarih[5:7]), int(tarih[8:10]))
                if gun.weekday() == 6:
                    gt = 'pazar'
                elif tarih in tatiller_map:
                    gt = 'tatil_tam' if tatiller_map[tarih]['tur'] == 'tam' else 'tatil_yarim'
                else:
                    gt = 'normal'

                eski_saat = eski['calisma_saati'] if eski else None
                if eski_saat != saat:
                    tum_degisen.append({
                        'personel': p_ad,
                        'personel_id': pid,
                        'tarih': tarih,
                        'eski_saat': eski_saat,
                        'yeni_saat': saat,
                        'gun_turu': gt,
                        'eski_gun_turu': eski['gun_turu'] if eski else None,
                    })

                conn.execute('''INSERT INTO puantaj (personel_id, tarih, gun_turu, calisma_saati)
                                VALUES (?,?,?,?)
                                ON CONFLICT(personel_id, tarih) DO UPDATE SET
                                gun_turu=excluded.gun_turu, calisma_saati=excluded.calisma_saati''',
                             (pid, tarih, gt, saat))

        # Tüm oturumu tek log satırı
        if tum_degisen or tum_silinen:
            degisen_personeller = list({d['personel'] for d in tum_degisen} |
                                       {d['personel'] for d in tum_silinen})
            ozet = {
                'personeller': degisen_personeller,
                'degistirilen_gun': len(tum_degisen),
                'silinen_gun': len(tum_silinen),
                'detay': tum_degisen,
                'silinen': tum_silinen,
            }
            conn.execute(
                'INSERT INTO degisiklik_log (tablo, kayit_id, islem, eski_deger, yeni_deger, tarih) VALUES (?,?,?,?,?,?)',
                ('puantaj', None, 'TOPLU', None, _json.dumps(ozet, ensure_ascii=False),
                 datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            )

        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'degisen': len(tum_degisen), 'silinen': len(tum_silinen)})
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'hata': str(e)}), 400

@app.route('/api/log/temizle', methods=['POST'])
def log_temizle():
    conn = get_db()
    conn.execute('DELETE FROM degisiklik_log')
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


def puantaj_sil(pid, tarih):
    conn = get_db()
    eski_r = conn.execute('SELECT * FROM puantaj WHERE personel_id=? AND tarih=?', (pid, tarih)).fetchone()
    if eski_r:
        log_yaz(conn, 'puantaj', eski_r['id'], 'DELETE', eski=dict(eski_r))
    conn.execute('DELETE FROM puantaj WHERE personel_id=? AND tarih=?', (pid, tarih))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# Avans
@app.route('/api/avans', methods=['POST'])
def avans_ekle():
    d = request.json
    conn = get_db()
    cur = conn.execute('INSERT INTO avans (personel_id, tutar, tarih, aciklama) VALUES (?,?,?,?)',
                 (d['personel_id'], d['tutar'], d['tarih'], d.get('aciklama', '')))
    log_yaz(conn, 'avans', cur.lastrowid, 'INSERT',
            yeni={'personel_id': d['personel_id'], 'tutar': d['tutar'],
                  'tarih': d['tarih'], 'aciklama': d.get('aciklama', '')})
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/avans/<int:pid>', methods=['GET'])
def avans_listesi(pid):
    yil = request.args.get('yil', datetime.now().year)
    ay = request.args.get('ay', datetime.now().month)
    ay_basi = f"{yil}-{int(ay):02d}-01"
    ay_sonu = f"{yil}-{int(ay):02d}-{calendar.monthrange(int(yil), int(ay))[1]:02d}"
    conn = get_db()
    rows = conn.execute('SELECT * FROM avans WHERE personel_id=? AND tarih>=? AND tarih<=? ORDER BY tarih',
                        (pid, ay_basi, ay_sonu)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/avans/<int:avans_id>', methods=['DELETE'])
def avans_sil(avans_id):
    conn = get_db()
    eski_r = conn.execute('SELECT * FROM avans WHERE id=?', (avans_id,)).fetchone()
    if eski_r:
        log_yaz(conn, 'avans', avans_id, 'DELETE', eski=dict(eski_r))
    conn.execute('DELETE FROM avans WHERE id=?', (avans_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# Hesaplama
@app.route('/api/hesapla/<int:pid>', methods=['GET'])
def maas_hesapla(pid):
    yil = int(request.args.get('yil', datetime.now().year))
    ay = int(request.args.get('ay', datetime.now().month))
    global_asgari = request.args.get('asgari', None)
    if global_asgari is not None:
        try:
            global_asgari = float(global_asgari)
        except ValueError:
            global_asgari = None
    sonuc = hesapla_maas(pid, yil, ay, global_asgari=global_asgari)
    if sonuc:
        return jsonify(sonuc)
    return jsonify({'hata': 'Personel bulunamadı'}), 404


TR_ALFABE = 'aAbBcCçÇdDeEfFgGğĞhHıIİijJkKlLmMnNoOöÖpPrRsSşŞtTuUüÜvVyYzZ'
def tr_sort_personel(rows, ad_key='ad', soyad_key='soyad'):
    def key(r):
        isim = (r[ad_key] + ' ' + r[soyad_key]).lower()
        return [TR_ALFABE.index(c) if c in TR_ALFABE else ord(c) for c in isim]
    return sorted(rows, key=key)

@app.route('/api/rapor', methods=['GET'])
def rapor():
    yil = int(request.args.get('yil', datetime.now().year))
    ay = int(request.args.get('ay', datetime.now().month))
    global_asgari = request.args.get('asgari', None)
    if global_asgari is not None:
        try:
            global_asgari = float(global_asgari)
        except ValueError:
            global_asgari = None
    conn = get_db()
    personeller = conn.execute('SELECT id, ad, soyad FROM personel WHERE aktif=1').fetchall()
    conn.close()
    sonuclar = [s for p in personeller if (s := hesapla_maas(p['id'], yil, ay, global_asgari=global_asgari))]
    tr_alfabe = 'aAbBcCçÇdDeEfFgGğĞhHıIİijJkKlLmMnNoOöÖpPrRsSşŞtTuUüÜvVyYzZ'
    def tr_key(s):
        isim = (s['personel']['ad'] + ' ' + s['personel']['soyad']).lower()
        return [tr_alfabe.index(c) if c in tr_alfabe else ord(c) for c in isim]
    sonuclar.sort(key=tr_key)
    return jsonify(sonuclar)

@app.route('/api/aylik_tablo', methods=['GET'])
def aylik_tablo():
    yil = int(request.args.get('yil', datetime.now().year))
    ay = int(request.args.get('ay', datetime.now().month))
    gun_sayisi = calendar.monthrange(yil, ay)[1]
    conn = get_db()
    personeller = tr_sort_personel(conn.execute('SELECT * FROM personel WHERE aktif=1').fetchall())
    tatiller = get_tatiller_dict(yil, ay)
    tablo = []
    for p in personeller:
        pid = p['id']
        ay_basi = f"{yil}-{ay:02d}-01"
        ay_sonu = f"{yil}-{ay:02d}-{gun_sayisi:02d}"
        kayitlar = conn.execute('SELECT tarih, calisma_saati, gun_turu FROM puantaj WHERE personel_id=? AND tarih>=? AND tarih<=?',
                                (pid, ay_basi, ay_sonu)).fetchall()
        kayit_map = {k['tarih']: k for k in kayitlar}
        hesap = hesapla_maas(pid, yil, ay)
        gunler = []
        for d in range(1, gun_sayisi + 1):
            tarih = f"{yil}-{ay:02d}-{d:02d}"
            k = kayit_map.get(tarih)
            gun_tarihi = date(yil, ay, d)
            gunler.append({
                'gun': d,
                'tarih': tarih,
                'pazar': gun_tarihi.weekday() == 6,
                'tatil': tatiller.get(tarih),
                'saat': k['calisma_saati'] if k else None,
                'gun_turu': k['gun_turu'] if k else None,
            })
        tablo.append({'personel': dict(p), 'gunler': gunler, 'ozet': hesap})
    conn.close()
    return jsonify({'yil': yil, 'ay': ay, 'gun_sayisi': gun_sayisi, 'tablo': tablo})

# ── XLSX Export ──────────────────────────────────────────────
@app.route('/api/xlsx_export', methods=['GET'])
def xlsx_export():
    yil = int(request.args.get('yil', datetime.now().year))
    ay = int(request.args.get('ay', datetime.now().month))
    kayit_yolu = request.args.get('kayit_yolu', None)  # diske kaydet modu
    gun_sayisi = calendar.monthrange(yil, ay)[1]
    ay_adlari = ['','Ocak','Şubat','Mart','Nisan','Mayıs','Haziran',
                 'Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık']

    conn = get_db()
    personeller = tr_sort_personel(conn.execute('SELECT * FROM personel WHERE aktif=1').fetchall())
    tatiller = get_tatiller_dict(yil, ay)
    tablo = []
    for p in personeller:
        pid = p['id']
        ay_basi = f"{yil}-{ay:02d}-01"
        ay_sonu = f"{yil}-{ay:02d}-{gun_sayisi:02d}"
        kayitlar = conn.execute('SELECT tarih, calisma_saati, gun_turu FROM puantaj WHERE personel_id=? AND tarih>=? AND tarih<=?',
                                (pid, ay_basi, ay_sonu)).fetchall()
        kayit_map = {k['tarih']: k for k in kayitlar}
        hesap = hesapla_maas(pid, yil, ay)
        tablo.append({'personel': dict(p), 'kayit_map': kayit_map, 'ozet': hesap})
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{ay_adlari[ay]} {yil}"

    # Renkler - Açık tema
    C_HEADER_BG  = "2563EB"   # Mavi header
    C_HEADER_FG  = "FFFFFF"
    C_PAZ_BG     = "FEF3C7"   # Sarı tonu - pazar
    C_PAZ_FG     = "B45309"
    C_TATIL_BG   = "DBEAFE"   # Açık mavi - tatil
    C_TATIL_FG   = "1D4ED8"
    C_NORMAL_BG  = "FFFFFF"   # Beyaz - normal
    C_NORMAL_FG  = "1E293B"
    C_FAZLA_FG   = "EA580C"   # Turuncu - fazla mesai
    C_TOPLAM_BG  = "EFF6FF"   # Çok açık mavi - toplam
    C_MAVI       = "2563EB"
    C_YESIL      = "16A34A"
    C_KIRMIZI    = "DC2626"

    def cell_style(ws, row, col, value, bg=None, fg="E4E8F5", bold=False,
                   align="center", num_fmt=None, border=True):
        c = ws.cell(row=row, column=col, value=value)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Calibri", color=fg, bold=bold, size=10)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        if num_fmt:
            c.number_format = num_fmt
        if border:
            thin = Side(style='thin', color="CBD5E1")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return c

    # ── Başlık ──
    baslik = f"PUANTAJ CETVELİ — {ay_adlari[ay].upper()} {yil}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + gun_sayisi + 2)
    c = ws.cell(row=1, column=1, value=baslik)
    c.fill = PatternFill("solid", fgColor=C_MAVI)
    c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Sütun başlıkları: Personel | 1 2 3 ... | Normal | Fazla ──
    COL_ISIM = 1
    COL_GUN_BASI = 2
    COL_TOPLAM = COL_GUN_BASI + gun_sayisi
    COL_FAZLA  = COL_TOPLAM + 1

    HEADER_ROW = 2
    ws.row_dimensions[HEADER_ROW].height = 36

    cell_style(ws, HEADER_ROW, COL_ISIM, "Personel",
               bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, align="left")
    ws.column_dimensions[get_column_letter(COL_ISIM)].width = 22

    gun_adlari = ["Pt","Sa","Ça","Pe","Cu","Ct","Pz"]
    for d in range(1, gun_sayisi + 1):
        col = COL_GUN_BASI + d - 1
        tarih = f"{yil}-{ay:02d}-{d:02d}"
        gun_tarihi = date(yil, ay, d)
        is_pazar = gun_tarihi.weekday() == 6
        is_tatil = tarih in tatiller
        gun_kisa = gun_adlari[gun_tarihi.weekday()]

        bg = C_PAZ_BG if is_pazar else (C_TATIL_BG if is_tatil else C_HEADER_BG)
        fg = C_PAZ_FG if is_pazar else (C_TATIL_FG if is_tatil else C_HEADER_FG)

        c = ws.cell(row=HEADER_ROW, column=col, value=d)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Calibri", color=fg, bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style='thin', color="CBD5E1")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Gün adı alt satır (3. satır)
        c2 = ws.cell(row=3, column=col, value=gun_kisa)
        c2.fill = PatternFill("solid", fgColor=bg)
        c2.font = Font(name="Calibri", color=fg, size=8)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.column_dimensions[get_column_letter(col)].width = 5.2

    cell_style(ws, HEADER_ROW, COL_TOPLAM, "Normal\nSaat",
               bg=C_HEADER_BG, fg=C_YESIL, bold=True)
    cell_style(ws, 3, COL_TOPLAM, "(s)", bg=C_HEADER_BG, fg=C_YESIL)
    ws.column_dimensions[get_column_letter(COL_TOPLAM)].width = 9

    cell_style(ws, HEADER_ROW, COL_FAZLA, "Fazla\nMesai",
               bg=C_HEADER_BG, fg=C_FAZLA_FG, bold=True)
    cell_style(ws, 3, COL_FAZLA, "(s)", bg=C_HEADER_BG, fg=C_FAZLA_FG)
    ws.column_dimensions[get_column_letter(COL_FAZLA)].width = 9

    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 16

    # Personel isim sütunu için HEADER_ROW ve 3. satır merge
    ws.merge_cells(start_row=HEADER_ROW, start_column=COL_ISIM, end_row=3, end_column=COL_ISIM)
    ws.merge_cells(start_row=HEADER_ROW, start_column=COL_TOPLAM, end_row=3, end_column=COL_TOPLAM)
    ws.merge_cells(start_row=HEADER_ROW, start_column=COL_FAZLA, end_row=3, end_column=COL_FAZLA)

    # ── Veri satırları ──
    DATA_ROW_START = 4
    for i, satir in enumerate(tablo):
        row = DATA_ROW_START + i
        ws.row_dimensions[row].height = 18
        p = satir['personel']
        ozet = satir['ozet']
        kayit_map = satir['kayit_map']

        isim = f"{p['ad']} {p['soyad']}"
        cell_style(ws, row, COL_ISIM, isim, bg=C_NORMAL_BG, fg=C_NORMAL_FG, bold=True, align="left")

        toplam_normal = 0.0
        toplam_fazla = 0.0

        for d in range(1, gun_sayisi + 1):
            col = COL_GUN_BASI + d - 1
            tarih_str = f"{yil}-{ay:02d}-{d:02d}"
            gun_tarihi = date(yil, ay, d)
            is_pazar = gun_tarihi.weekday() == 6
            tatil = tatiller.get(tarih_str)
            k = kayit_map.get(tarih_str)
            saat = k['calisma_saati'] if k else None

            if is_pazar:
                bg = C_PAZ_BG
                if saat is None:
                    val, fg = "", C_PAZ_FG
                elif saat == 0:
                    val, fg = 7.5, C_PAZ_FG
                    toplam_normal += 7.5
                else:
                    val, fg = saat, C_PAZ_FG
                    toplam_normal += min(saat, TAVAN_SAAT)
                    if saat > TAVAN_SAAT:
                        toplam_fazla += saat - TAVAN_SAAT
            elif tatil:
                bg = C_TATIL_BG
                baz = TAVAN_SAAT if tatil['tur'] == 'tam' else TAVAN_SAAT / 2
                if saat is None:
                    # Boş → hiç sayılmaz, hücre boş göster
                    val, fg = "", C_TATIL_FG
                else:
                    val, fg = saat, C_TATIL_FG
                    # Her durumda 7.5 normale eklenir, baz üstü fazla mesai
                    toplam_normal += TAVAN_SAAT
                    if saat > baz:
                        toplam_fazla += saat - baz
            else:
                bg = C_NORMAL_BG
                if saat is None:
                    val, fg = "", C_NORMAL_FG
                elif saat > TAVAN_SAAT:
                    val = saat
                    fg = C_FAZLA_FG
                    toplam_normal += TAVAN_SAAT
                    toplam_fazla += saat - TAVAN_SAAT
                else:
                    val, fg = saat, C_NORMAL_FG
                    toplam_normal += saat

            c = ws.cell(row=row, column=col, value=val if val != "" else None)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Calibri", color=fg, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")
            thin = Side(style='thin', color="CBD5E1")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if isinstance(val, float) and val != "":
                c.number_format = '0.0'

        cell_style(ws, row, COL_TOPLAM, round(toplam_normal, 2),
                   bg=C_TOPLAM_BG, fg=C_YESIL, bold=True, num_fmt='0.0')
        cell_style(ws, row, COL_FAZLA, round(toplam_fazla, 2) if toplam_fazla > 0 else None,
                   bg=C_TOPLAM_BG, fg=C_FAZLA_FG if toplam_fazla > 0 else "94A3B8", bold=toplam_fazla > 0, num_fmt='0.0')

    # ── Toplam satırı ──
    TOPLAM_ROW = DATA_ROW_START + len(tablo)
    ws.row_dimensions[TOPLAM_ROW].height = 20
    cell_style(ws, TOPLAM_ROW, COL_ISIM, "TOPLAM", bg="DBEAFE", fg=C_MAVI, bold=True, align="left")
    for d in range(1, gun_sayisi + 1):
        col = COL_GUN_BASI + d - 1
        cell_style(ws, TOPLAM_ROW, col, "", bg="DBEAFE", fg=C_NORMAL_FG)

    # Toplam normal ve fazla hesapla
    t_normal = sum(s['ozet']['normal_saat'] + s['ozet']['yillik_izin_saat'] for s in tablo if s['ozet'])
    t_fazla  = sum(s['ozet']['fazla_mesai_saat'] for s in tablo if s['ozet'])
    cell_style(ws, TOPLAM_ROW, COL_TOPLAM, round(t_normal, 2), bg="DBEAFE", fg=C_YESIL, bold=True, num_fmt='0.0')
    cell_style(ws, TOPLAM_ROW, COL_FAZLA,  round(t_fazla, 2),  bg="DBEAFE", fg=C_FAZLA_FG, bold=True, num_fmt='0.0')

    # Dondur: başlık + isim sütunu
    ws.freeze_panes = ws.cell(row=DATA_ROW_START, column=COL_GUN_BASI)

    dosya_adi = f"puantaj_{yil}_{ay:02d}.xlsx"
    if kayit_yolu:
        # Diske kaydet modu (PyWebView) — kayit_yolu tam dosya yolunu içerir
        tam_yol = kayit_yolu if kayit_yolu.endswith('.xlsx') else os.path.join(kayit_yolu, dosya_adi)
        wb.save(tam_yol)
        return jsonify({'ok': True, 'yol': tam_yol})
    else:
        # Tarayıcı indirme modu
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=dosya_adi,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Değişiklik Geçmişi
@app.route('/api/log', methods=['GET'])
def log_listesi():
    limit = int(request.args.get('limit', 200))
    tablo = request.args.get('tablo', '')
    conn = get_db()
    if tablo == 'TOPLU':
        rows = conn.execute(
            "SELECT * FROM degisiklik_log WHERE islem='TOPLU' ORDER BY id DESC LIMIT ?",
            (limit,)).fetchall()
    elif tablo:
        rows = conn.execute(
            "SELECT * FROM degisiklik_log WHERE tablo=? AND islem!='TOPLU' ORDER BY id DESC LIMIT ?",
            (tablo, limit)).fetchall()
    else:
        rows = conn.execute(
            'SELECT * FROM degisiklik_log ORDER BY id DESC LIMIT ?',
            (limit,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/log/geri/<int:log_id>', methods=['POST'])
def log_geri_al(log_id):
    import json
    conn = get_db()
    log = conn.execute('SELECT * FROM degisiklik_log WHERE id=?', (log_id,)).fetchone()
    if not log:
        conn.close()
        return jsonify({'ok': False, 'hata': 'Log kaydı bulunamadı'}), 404

    tablo = log['tablo']
    islem = log['islem']
    eski = json.loads(log['eski_deger']) if log['eski_deger'] else None
    kayit_id = log['kayit_id']

    try:
        if islem == 'DELETE' and eski:
            # Silinen kaydı geri ekle
            if tablo == 'puantaj':
                conn.execute('''INSERT OR REPLACE INTO puantaj
                    (personel_id, tarih, gun_turu, calisma_saati, notlar)
                    VALUES (?,?,?,?,?)''',
                    (eski['personel_id'], eski['tarih'], eski['gun_turu'],
                     eski['calisma_saati'], eski.get('notlar', '')))
                log_yaz(conn, tablo, kayit_id, 'INSERT', yeni=eski)
            elif tablo == 'avans':
                conn.execute('INSERT OR IGNORE INTO avans (id,personel_id,tutar,tarih,aciklama) VALUES (?,?,?,?,?)',
                    (eski['id'], eski['personel_id'], eski['tutar'], eski['tarih'], eski.get('aciklama','')))
                log_yaz(conn, tablo, kayit_id, 'INSERT', yeni=eski)
            elif tablo == 'tatil':
                conn.execute('INSERT OR IGNORE INTO tatil (id,tarih,ad,tur) VALUES (?,?,?,?)',
                    (eski['id'], eski['tarih'], eski['ad'], eski.get('tur','tam')))
                log_yaz(conn, tablo, kayit_id, 'INSERT', yeni=eski)
            elif tablo == 'personel':
                conn.execute('UPDATE personel SET aktif=1 WHERE id=?', (kayit_id,))
                log_yaz(conn, tablo, kayit_id, 'UPDATE', yeni={'aktif': 1})
        elif islem in ('INSERT', 'UPDATE') and eski:
            # Güncellenen/eklenen kaydı eski haline döndür
            if tablo == 'puantaj':
                conn.execute('''INSERT OR REPLACE INTO puantaj
                    (personel_id, tarih, gun_turu, calisma_saati, notlar)
                    VALUES (?,?,?,?,?)''',
                    (eski['personel_id'], eski['tarih'], eski['gun_turu'],
                     eski['calisma_saati'], eski.get('notlar', '')))
                log_yaz(conn, tablo, kayit_id, 'UPDATE', yeni=eski)
            elif tablo == 'maas_gecmis':
                conn.execute('UPDATE maas_gecmis SET yil=?,ay=?,sabit_maas=? WHERE id=?',
                    (eski['yil'], eski.get('ay'), eski['sabit_maas'], kayit_id))
                log_yaz(conn, tablo, kayit_id, 'UPDATE', yeni=eski)
            elif tablo == 'personel':
                conn.execute('UPDATE personel SET ad=?,soyad=?,tc=? WHERE id=?',
                    (eski['ad'], eski['soyad'], eski['tc'], kayit_id))
                log_yaz(conn, tablo, kayit_id, 'UPDATE', yeni=eski)
        elif islem == 'TOPLU':
            import json as _j
            ozet = _j.loads(log['yeni_deger']) if log['yeni_deger'] else {}
            # Değiştirilen günleri eski saatlerine döndür
            for d in ozet.get('detay', []):
                pid = d['personel_id']
                if d['eski_saat'] is None:
                    conn.execute('DELETE FROM puantaj WHERE personel_id=? AND tarih=?', (pid, d['tarih']))
                else:
                    conn.execute('''INSERT INTO puantaj (personel_id, tarih, gun_turu, calisma_saati)
                        VALUES (?,?,?,?)
                        ON CONFLICT(personel_id, tarih) DO UPDATE SET
                        gun_turu=excluded.gun_turu, calisma_saati=excluded.calisma_saati''',
                        (pid, d['tarih'], d.get('eski_gun_turu') or 'normal', d['eski_saat']))
            # Silinenleri geri ekle
            for d in ozet.get('silinen', []):
                conn.execute('''INSERT OR IGNORE INTO puantaj
                    (personel_id, tarih, gun_turu, calisma_saati, notlar)
                    VALUES (?,?,?,?,?)''',
                    (d['personel_id'], d['tarih'], d['gun_turu'],
                     d['calisma_saati'], d.get('notlar', '')))
            log_yaz(conn, 'puantaj', None, 'TOPLU', yeni={'geri_alindi': True, 'log_id': log_id})
            # Yeni eklenen kaydı sil
            if tablo == 'puantaj':
                conn.execute('DELETE FROM puantaj WHERE id=?', (kayit_id,))
                log_yaz(conn, tablo, kayit_id, 'DELETE')
            elif tablo == 'avans':
                conn.execute('DELETE FROM avans WHERE id=?', (kayit_id,))
                log_yaz(conn, tablo, kayit_id, 'DELETE')
            elif tablo == 'tatil':
                conn.execute('DELETE FROM tatil WHERE id=?', (kayit_id,))
                log_yaz(conn, tablo, kayit_id, 'DELETE')
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'hata': str(e)}), 500

# Ayarlar
@app.route('/api/ayarlar/<anahtar>', methods=['GET'])
def ayar_get(anahtar):
    conn = get_db()
    r = conn.execute('SELECT deger FROM ayarlar WHERE anahtar=?', (anahtar,)).fetchone()
    conn.close()
    return jsonify({'deger': r['deger'] if r else None})

@app.route('/api/ayarlar/<anahtar>', methods=['POST'])
def ayar_set(anahtar):
    d = request.json
    deger = d.get('deger', '')
    conn = get_db()
    if deger == '' or deger is None:
        conn.execute('DELETE FROM ayarlar WHERE anahtar=?', (anahtar,))
    else:
        conn.execute('''INSERT INTO ayarlar (anahtar, deger) VALUES (?,?)
                        ON CONFLICT(anahtar) DO UPDATE SET deger=excluded.deger''',
                     (anahtar, str(deger)))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# Yedek
@app.route('/api/yedek', methods=['POST'])
def yedek_olustur():
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)
    zaman = datetime.now().strftime('%Y%m%d_%H%M%S')
    hedef = os.path.join(BACKUP_DIR, f'puantaj_{zaman}.db')
    shutil.copy2(DB_PATH, hedef)
    return jsonify({'ok': True, 'dosya': os.path.basename(hedef)})

@app.route('/api/yedek', methods=['GET'])
def yedek_listesi():
    if not os.path.exists(BACKUP_DIR):
        return jsonify([])
    dosyalar = []
    for f in sorted(os.listdir(BACKUP_DIR), reverse=True):
        if f.endswith('.db'):
            yol = os.path.join(BACKUP_DIR, f)
            dosyalar.append({
                'ad': f,
                'boyut': round(os.path.getsize(yol) / 1024, 1),
                'tarih': datetime.fromtimestamp(os.path.getmtime(yol)).strftime('%d.%m.%Y %H:%M')
            })
    return jsonify(dosyalar)

@app.route('/api/yedek/geri/<dosya>', methods=['POST'])
def yedek_geri(dosya):
    yol = os.path.join(BACKUP_DIR, dosya)
    if not os.path.exists(yol):
        return jsonify({'ok': False, 'hata': 'Dosya bulunamadı'}), 404
    zaman = datetime.now().strftime('%Y%m%d_%H%M%S')
    shutil.copy2(DB_PATH, os.path.join(BACKUP_DIR, f'onceki_{zaman}.db'))
    shutil.copy2(yol, DB_PATH)
    return jsonify({'ok': True})

@app.route('/api/yedek/sil/<dosya>', methods=['DELETE'])
def yedek_sil(dosya):
    yol = os.path.join(BACKUP_DIR, dosya)
    if not os.path.exists(yol):
        return jsonify({'ok': False, 'hata': 'Dosya bulunamadı'}), 404
    os.remove(yol)
    return jsonify({'ok': True})

if __name__ == '__main__':
    # ── TEK ÖRNEK KONTROLÜ (Windows Mutex) ──────────────────
    import ctypes
    _MUTEX_ADI = 'PuantajSistemi_TekOrnek_Mutex'
    _mutex = ctypes.windll.kernel32.CreateMutexW(None, False, _MUTEX_ADI)
    _hata = ctypes.windll.kernel32.GetLastError()
    if _hata == 183:  # ERROR_ALREADY_EXISTS
        # Zaten açık — mevcut pencereyi öne getir ve kapat
        import ctypes.wintypes
        HWND = ctypes.windll.user32.FindWindowW(None, 'Puantaj Sistemi')
        if HWND:
            ctypes.windll.user32.ShowWindow(HWND, 9)   # SW_RESTORE
            ctypes.windll.user32.SetForegroundWindow(HWND)
        import sys
        sys.exit(0)
    # ────────────────────────────────────────────────────────

    init_db()
    import threading, time

    def start_flask():
        app.run(debug=False, port=5000, use_reloader=False)

    t = threading.Thread(target=start_flask, daemon=True)
    t.start()
    time.sleep(1.5)

    try:
        import webview

        class PuantajAPI:
            def klasor_sec(self):
                result = webview.windows[0].create_file_dialog(
                    webview.FOLDER_DIALOG
                )
                return list(result) if result else []

            def dosya_kaydet_sec(self, varsayilan_ad):
                result = webview.windows[0].create_file_dialog(
                    webview.SAVE_DIALOG,
                    save_filename=varsayilan_ad,
                    file_types=('Excel Dosyası (*.xlsx)',)
                )
                return result if result else None

        api_obj = PuantajAPI()
        window = webview.create_window(
            title='Puantaj Sistemi',
            url='http://localhost:5000',
            width=1280,
            height=800,
            min_size=(800, 600),
            resizable=True,
            text_select=True,
            js_api=api_obj,
        )
        webview.start(debug=False)
    except Exception:
        import subprocess
        subprocess.Popen(['cmd', '/c', 'start', '', 'http://localhost:5000'],
                        shell=False, creationflags=0x08000000)
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            pass